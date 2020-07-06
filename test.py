# import the necessary packages
from openpyxl import *
from openpyxl.styles import Side , Border , PatternFill , Alignment , Font
import os , subprocess
import win32ctypes
import pkg_resources.py2_warn

# opening the targeted form and make it protected to limit mistakes
wb = load_workbook('./رصيد.xlsx')
work_sheet = wb.active
work_sheet.protection.sheet = True

# create a name and get the current directory of this created file
form_title = "one"
the_current_directory = subprocess.getoutput("cd")


def create_a_new_excel_file():

    def create_the_workbook():
        # create a file if it was not already there
        li = subprocess.getoutput("dir")
        if f"{form_title}.xlsx" in li:
            new_work_book = load_workbook(f'./{form_title}.xlsx')
        else:
            new_work_book = Workbook()
        return new_work_book

    new_work_book = create_the_workbook()
    # change the name of the first sheet to (1) :)
    sh = new_work_book.active
    if sh.title == "Sheet":
        sh.title = '(1)'
    # three variables to help create dynamic sheets number
    number_of_sheets = len(new_work_book.worksheets)
    flag = False
    count_flag = False

    def dynamic_sheets():
        nonlocal number_of_sheets
        nonlocal flag
        nonlocal count_flag

        # if the sheets are full create the next 10 sheets
        if new_work_book[f'({number_of_sheets})']['c2'].value != None or number_of_sheets == 1:
            if number_of_sheets == 1 and new_work_book[f'({number_of_sheets})']['c2'].value == None:
                count_flag = True
            if new_work_book[f'({number_of_sheets})']['c2'].value != None:
                count = number_of_sheets
                while count < number_of_sheets + 10 :
                    new_work_book.create_sheet(title=f'({count + 1})')
                    count += 1
                number_of_sheets += 10
            flag = True

    dynamic_sheets()

    def format_the_cells():
        sheet.column_dimensions['c'].width = 26
        sheet.column_dimensions['a'].width = 13.5
        sheet.column_dimensions['d'].width = 13
        sheet.column_dimensions['e'].width = 10
        sheet.column_dimensions['f'].width = 10.5
        sheet.column_dimensions['b'].width = 8
        for i in range(1, 100):
            if i == 3 or i ==4 :
                sheet.row_dimensions[i].height = 18
                continue
            sheet.row_dimensions[i].height = 26.5
        sheet.freeze_panes = "A5"
        cells = sheet['A5:F100']
        t = Side('thin')
        border_style = Border(left=t, right=t, top=t, bottom=t)
        for col in cells:
            for cell in col:
                cell.border = border_style

    def merge_cells_for_the_same_format():
        sheet.merge_cells('a3:a4')
        sheet.merge_cells('b3:b4')
        sheet.merge_cells('c3:c4')
        sheet.merge_cells('d3:e3')
        sheet.merge_cells('f3:f4')

    def color_the_background():
        double = Side(border_style='double', color='000000')
        fill = PatternFill(patternType='solid',fgColor='00FF1493' , bgColor='00FFFFFF')
        c = sheet['c2']
        c.fill = PatternFill(patternType='solid',fgColor='00FFFF00' , bgColor='00FFFFFF')
        cells_list = ['a3' , 'b3' , 'c3' , 'd3' , 'd4' , 'e4' , 'f3']
        for cell in cells_list:
            ce = sheet[cell]
            ce.fill = fill
            ce.border = Border(left=double , right=double , top = double , bottom= double)

    def align_cells_in_center():
        cells = sheet['a1' : 'f100']
        for r in cells:
            for cel in r:
                cel.alignment = Alignment(horizontal='center' , vertical='center')

    def set_the_fonts():
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                c = sheet.cell(row= i + 1 , column= j + 1)
                c.font = Font(name= "Traditional Arabic",size= 14 , bold=True , italic=True)

    def set_income_import_export_formulas():
        for i in range(5,100):
            val = f'=+IF(SUM($E$1:E{i})>SUM($D$1:D{i}),SUM($E$1:E{i})-SUM($D$1:D{i}),"0.00")'
            sheet[f'F{i}'].value = val
        sheet['E100'].value = '=SUM(E5:E99)'
        sheet['D100'].value = '=SUM(D5:D99)'

    # using the flag to know if a new sheets have been built and if so, give them the format of the module suggested
    if flag:
        if count_flag:
            count = 1
        else:
            count = number_of_sheets - 10
        while count <= number_of_sheets:
            sheet = new_work_book[f'({count})']
            format_the_cells()
            color_the_background()

            sheet['a1'].value = 'سعة الكرتونة'
            sheet['b1'].value = 'الماركة'
            sheet['f1'].value = 'المقاس'
            merge_cells_for_the_same_format()
            sheet['a3'].value = 'التاريخ'
            sheet['b3'].value = 'رقم المستند'
            sheet['c3'].value = 'البيان'
            sheet['d3'].value = 'الحركة'
            sheet['d4'].value = 'صادر'
            sheet['e4'].value = 'وارد'
            sheet['f3'].value = 'الرصيد'

            set_the_fonts()
            align_cells_in_center()
            set_income_import_export_formulas()
            count += 1

    new_work_book.save(f'./{form_title}.xlsx')


def variables_collecting(work_sh):
    type_name = 'c2'
    size = 'f2'
    brand = 'b2'
    box_size = 'a2'
    income = 'f99'
    imported = 'e100'
    exported = 'd100'
    info_list = [type_name , box_size , brand , size , income , imported , exported]
    return info_list


create_a_new_excel_file()


def open_the_new_excel_file():
    p = f'{the_current_directory}/{form_title}.xlsx'
    os.system(p)


open_the_new_excel_file()
wb1 = load_workbook(f'./{form_title}.xlsx')

# linking the sheets to رصيد sheet to help doing inventory work
def link_the_new_excel_file():
    for cur in range(1, len(wb1.worksheets) + 1):
        ws1 = wb1[f'({str(cur)})']
        var_collected = variables_collecting(ws1)
        count = 0
        for le in "DEFGHIJ":
            work_sheet[f'{le}{cur + 3}'].value = f"='{the_current_directory}\[{form_title}.xlsx]({str(cur)})'!{var_collected[count]}"
            count += 1
    wb.save('رصيد.xlsx')


link_the_new_excel_file()


